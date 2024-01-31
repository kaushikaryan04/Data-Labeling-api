from rest_framework.decorators import api_view
from .serializers import TagSerializer , FileSerializer
from rest_framework.parsers import FileUploadParser
from django.http import FileResponse
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Statement , Tag
import pandas as pd
import os 
from openpyxl import load_workbook
from core import settings
from django.core.cache import cache

class FileUploadView(APIView):
    parser_classes = [FileUploadParser]
    def post(self , request , *args , **kwargs ): 
        file = request.FILES.get('file')
        if file :
            destination_path = os.path.join(settings.BASE_DIR ,'files', file.name)
            file_format = is_file_xlsx(file.name)
            if file_format == "xlsx" : 
                df = pd.read_excel(file,engine='openpyxl')
                print(df)
                df.to_excel(destination_path , index = False )
                statements = df['Statement']
            elif file_format == "csv" :

                df = pd.read_csv(file , header= None ) 
                # print(df)
                df = df.iloc[3:-1]
                statements = []

                for i in range(len(df)) :
                    original_string = df.iloc[i , 0]
                    statements.append(original_string)
                print(df)
                df.to_csv(destination_path,index= False)

            else :
                return Response({
                    "status" : "File upload failed check server logs"
                })

            

            for statement in statements:
                s = Statement.objects.create(
                    text = str(statement)
                )
                s.save()
            return Response({
                "status" : "file uploadSuccessfull" 
            })
        return Response({
            "status" : "File upload failed check server logs"
        })

class TagStatementCSView(APIView):
    def post(self , request , *args ,**kwargs) :
        data = request.data.get('data' , [])
        file_name = request.data.get("filename")
        serialized = TagSerializer(data = data , many =  True)
        if serialized.is_valid() :
            serialized_data = serialized.data 
            data_row = []
            for data in serialized_data :
                temp = {
                    "Statement" : data["statement"],
                    "Aspect" : data["aspect"],
                    "Sentiment" : data["sentiment"]
                }
                data_row.append(temp)
                tag = Tag.objects.create(
                    statement = Statement.objects.get(text = data["statement"]),
                    aspect = data["aspect"],
                    sentiment = data["sentiment"]
                )
                tag.save()
        else :
            return Response({
                "Error":"Request format for data not valid"
            })
        df_to_append = pd.DataFrame(data_row)
        csv_file_path = os.path.join(settings.BASE_DIR , "files" , f"new{file_name}")
        if os.path.exists(csv_file_path):
            existing_df = pd.read_csv(csv_file_path)
            df_combined = pd.concat([existing_df, df_to_append], ignore_index=True)
        else :
            df_combined = df_to_append
        df_combined.to_csv(csv_file_path, index=False)
        return Response({
            "File saved" : "Success"
        })





class TagStatementView(APIView) :
    def post(self , request ,*args , **kwargs) :
        data = request.data.get('data', [])
        file_name = request.data.get("filename")
        serialized = TagSerializer(data = data , many = True)

        if serialized.is_valid() :
            serialized_data = serialized.data 
            data_row = []
            for data in serialized_data :
                temp = [data["statement"],data["aspect"],data["sentiment"]]
                data_row.append(temp)
                tag = Tag.objects.create(
                    statement = Statement.objects.get(text = data["statement"]),
                    aspect = data["aspect"],
                    sentiment = data["sentiment"]
                )
                tag.save()
        else :
            return Response({
                "Error":"Request format for data not valid"
            })
        headings_for_file = ["Statement" , "Aspect" , "Sentiment"]
        # try : 
        file_path = os.path.join(settings.BASE_DIR, "files" ,file_name )
        workbook = load_workbook(file_path)
        sheet_exists  = "tags" in workbook.sheetnames
        df = pd.DataFrame(data_row, columns = headings_for_file)
        if sheet_exists :
            # print("in tags sheet")
            writer =   pd.ExcelWriter(file_path,  engine = 'openpyxl' , mode = "a" , if_sheet_exists= "overlay")
            df.to_excel(writer ,sheet_name="tags",index = False ,  header= False ,startrow=len(pd.read_excel(file_path , "tags"))+1 )
            writer._save()
            return Response({"Success":"Tags added successfully"})
        else :
            print(file_path)
            with pd.ExcelWriter(file_path,engine="openpyxl" , mode = "a") as writer :
                df.to_excel(writer , "tags" , index= False , header=True)
            return Response({"Success":"Tags added successfully"})


class DownloadTaggedFile(APIView ) :
    def get(self , request , *args , **kwargs) :
        file_name = str(request.GET.get("file_name"))
        if is_file_xlsx(file_name) == "csv":
            file_name = f"new{file_name}"
        file_path = os.path.join(settings.BASE_DIR , "files" , file_name)
        if os.path.exists(file_path) :
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response 

        else :
            return Response({
                "Error" : "First upload file with statements"
            })

@api_view(["GET"])
def get_all_aspects(request) :
    statement = request.GET.get("statement")
    key = "".join(statement.split())
    data = cache.get(key)
    if data is not None :
        return Response(data)
    queryset = Tag.objects.filter(statement = Statement.objects.get(text = statement))
    serializer = TagSerializer(queryset , many = True)
    data = serializer.data
    cache.set(key , data , timeout = 3600)
    return Response(data)



def is_file_xlsx(filename):
    if ".xlsx" in filename :
        return "xlsx"
    elif ".csv" in filename:
        return "csv" 
    return None 